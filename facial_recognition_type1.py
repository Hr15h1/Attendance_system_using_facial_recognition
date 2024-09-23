import os
import cv2
from deepface import DeepFace

from openpyxl import Workbook, load_workbook

from datetime import datetime, timedelta
import csv
import time


s = 1
source = cv2.VideoCapture(s) 
year = datetime.now().year
title = datetime.now().strftime("%B_%Y")






def stop_camera(window_name):

    # Release the video source and destroy the window
    source.release()
    cv2.destroyWindow(window_name)

def start_camera():
    alive = True
    window_name = "Attendance System"
    # Create a window to display the video feed
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)


    # Specify the CSV file name
    csv_file = "students.csv"

    # Initialize an empty dictionary
    students_dict = {}

    # Read the data from the CSV file into the dictionary
    with open(csv_file, mode='r', newline='') as file:
        reader = csv.DictReader(file)
        for row in reader:
            # Assuming the first column is the key and the second column is the value
            key = int(row["ID"])  # Convert the key to an integer
            value = row["Name"]
            students_dict[key] = value


    present_time = datetime.now().strftime("%I:%M:%S")
    year = datetime.now().year
    title = datetime.now().strftime("%B_%Y")
    today = datetime.now().strftime("%d-%m-%Y")

    #Try to load the attendance workbook
    try:
        wb = load_workbook(f"attendance_{datetime.now().year}.xlsx")
    #If it doesn't exist, create a new workbook
    except FileNotFoundError:
        wb = Workbook()
        wb.save(f"attendance_{datetime.now().year}.xlsx")
        wb = load_workbook(f"attendance_{datetime.now().year}.xlsx")

    #Check if the sheet for the current month exists, if not create it
    try:
        ws = wb[title]
    except KeyError:
        ws = wb.create_sheet(title, 0)
        first_day_of_month = datetime(year, datetime.now().month, 1)
        next_month = first_day_of_month.replace(month=first_day_of_month.month % 12 + 1, day=1)
        num_days = (next_month - first_day_of_month).days
        ws.cell(row=1, column=1, value="")
        ws.cell(row=1, column=2, value="")
        for i in range(1, 21):
            ws.cell(row=1, column=i + 2, value=i)

        for day in range(1, num_days + 1):
            date_str = (first_day_of_month + timedelta(days=day - 1)).strftime("%d-%m-%Y")
            ws.cell(row=day * 2, column=1, value=date_str)
            ws.cell(row=day * 2, column=2, value="am")
            ws.cell(row=day * 2 + 1, column=2, value="pm")
        wb.save(f"attendance_{year}.xlsx")




    
    # start_time = time.time()
    while alive:
        has_frame, frame = source.read()
        if not has_frame:
            break
        frame = cv2.flip(frame, 1)
        # frame_bw = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        # resized_frame = cv2.resize(frame, (320, 240))
        # current_time = time.time()
        
        # interval = 5

        # elapsed_time = current_time - start_time

        # if elapsed_time >= interval:
        #     has_frame, frame = source.read()
        #     if not has_frame:
        #         break
        #     frame = cv2.flip(frame, 1)

        #     resized_frame = cv2.resize(frame, (640, 480))

        try:
            # Perform face recognition by passing each frame to the models
            result1 = DeepFace.find(frame, db_path = "./database", enforce_detection = False, model_name = "Dlib", detector_backend = "dlib", align = True, distance_metric = "euclidean", anti_spoofing = True)
            result2 = DeepFace.find(frame, db_path = "./database", enforce_detection = False, model_name = "VGG-Face", detector_backend = "ssd", align = False, distance_metric = "euclidean_l2", anti_spoofing = True)
            result3 = DeepFace.find(frame, db_path = "./database", enforce_detection = False, model_name = "ArcFace", detector_backend = "yunet", align = True, distance_metric = "euclidean_l2", anti_spoofing = True)
            result4 = DeepFace.find(frame, db_path = "./database", enforce_detection = False, model_name = "GhostFaceNet", detector_backend = "opencv", align = False, distance_metric = "cosine", anti_spoofing = True)
            result5 = DeepFace.find(frame, db_path = "./database", enforce_detection = False, model_name = "SFace", detector_backend = "ssd", align = True, distance_metric = "euclidean", anti_spoofing = True)
            result6 = DeepFace.find(frame, db_path = "./database", enforce_detection = False, model_name = "Facenet512", detector_backend = "retinaface", align = True, distance_metric = "cosine", anti_spoofing = True)
            result7 = DeepFace.find(frame, db_path = "./database", enforce_detection = False, model_name = "Facenet", detector_backend = "yunet", align = True, distance_metric = "euclidean_l2", anti_spoofing = True)

            if len(result1[0]['identity']) > 0:
                
                n1 = result1[0]['identity'][0].split('\\')[1]
            else:
                n1 = 'empty'
            if len(result2[0]['identity']) > 0:
                
                n2 = result2[0]['identity'][0].split('\\')[1]
            else:
                n2 = 'empty'
            if len(result3[0]['identity']) > 0:
                
                n3 = result3[0]['identity'][0].split('\\')[1]
            else:
                n3 = 'empty'
            if len(result4[0]['identity']) > 0:
                
                n4 = result4[0]['identity'][0].split('\\')[1]
            else:
                n4 = 'empty'
            if len(result5[0]['identity']) > 0:
                
                n5 = result5[0]['identity'][0].split('\\')[1]
            else:
                n5 = 'empty'
            if len(result6[0]['identity']) > 0:
                
                n6 = result6[0]['identity'][0].split('\\')[1]
            else:
                n6 = 'empty'
            if len(result7[0]['identity']) > 0:
                
                n7 = result7[0]['identity'][0].split('\\')[1]
            else:
                n7 = 'empty'

        

            n = [n1, n2, n3, n4, n5, n6, n7] #list of names generated by the models

            name = max(set(n), key = n.count) #get the most common name from the list n


            # If the name is empty, assign the first non-empty name from list n
            if name == 'empty':
                for names in n:
                    if names != 'empty':
                        name = names


            if len(result1[0]['identity']) > 0:
                # name = n1
                # Get the bounding box coordinates
                xmin = int(result1[0]['source_x'][0])
                ymin = int(result1[0]['source_y'][0])
                w = result1[0]['source_w'][0]
                h = result1[0]['source_h'][0]

                xmax = int(xmin + w)
                ymax = int(ymin + h)

                cv2.rectangle(frame, (xmin, ymin), (xmax, ymax), (0, 255, 0), 2)
                cv2.rectangle(frame, (xmin, ymin - 25), (xmax, ymin),(255, 255, 255), -1)
                cv2.putText(frame, name, (xmin, ymin), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0),2, cv2.LINE_AA)

                #Mark attendance

                for key in students_dict.keys():
                    if students_dict[key] == name:
                        roll_no = key
                        break
                
                for col in ws.iter_cols(min_row = 1, max_col = 1, max_row = ws.max_row):
                    for cell in col:
                        if cell.value == today:
                            if datetime.now().hour < 12:
                                target_cell = ws.cell(row = cell.row, column = roll_no + 2)
                            else:
                                target_cell = ws.cell(row = cell.row + 1, column = roll_no + 2)                                

                            if target_cell.value is None:
                                target_cell.value = present_time
                                wb.save(f"attendance_{year}.xlsx")
                                print(f"Attendance marked for {name} at {present_time}")
                            else:
                                print(f"Attendance already marked for {name} at {target_cell.value}")
                            break
            



        except ValueError as e:
            print(f"Error: {e}")
            #Perform spoof detection if any of the models in the try block detects anti_spoofing
            spoof_result = DeepFace.extract_faces(frame, detector_backend = "opencv", enforce_detection = False, align = False, anti_spoofing = True)
            if len(spoof_result) > 0:
                x_min = spoof_result[0]['facial_area']['x']
                y_min = spoof_result[0]['facial_area']['y']
                x_max = x_min + spoof_result[0]['facial_area']['w']
                y_max = y_min + spoof_result[0]['facial_area']['h']
                cv2.rectangle(frame, (x_min, y_min), (x_max, y_max), (0, 0, 255), 2)
                cv2.rectangle(frame, (x_min, y_min - 25), (x_max, y_min), (255, 255, 255), -1)
                cv2.putText(frame, "Spoofing Detected", (x_min, y_min), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2, cv2.LINE_AA)
        
        # Display the frame
        cv2.imshow(window_name, frame)
        cv2.resizeWindow(window_name, frame.shape[1], frame.shape[0])
        # Check for key press to exit
        key = cv2.waitKey(1)
        if key == ord('q') or key == ord('Q') or key == 27:
            alive = False

    stop_camera(window_name)

start_camera()
