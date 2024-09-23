import os
import time
import csv
import cv2
import numpy as np
import pickle
from datetime import datetime
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta


face_detect = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
c = 1
video = cv2.VideoCapture(c)

def get_name_from_user():
    name = input("Enter your name: ")
    return name


def stop_camera(window_name):
    video.release()
    cv2.destroyWindow(window_name)
    return




def start_camera(purpose, no_of_faces = 70):
    
  
    alive = True
    window_name = "Add faces"
    cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)

    if not video.isOpened():
        print("Error: Could not open video.")
        return
    
    if purpose == "add":
        i = 0
        faces_data = []
        name = get_name_from_user()

        while alive:
            ret, frame = video.read()
            if not ret:
                print("Error: Could not read frame.")
                break

            frame = cv2.flip(frame, 1)

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY) 

            faces = face_detect.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4)

            for (x, y, w, h) in faces:
                crop_face = frame[y:y+h, x:x+w, :]
                resized_face = cv2.resize(crop_face, (100, 100))
                if len(faces_data) <= no_of_faces and i % 10 == 0:
                    faces_data.append(resized_face)
                i += 1
                cv2.putText(frame, str(len(faces_data)), (50, 50), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 1)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 1)
                


            cv2.imshow(window_name, frame)
            cv2.resizeWindow(window_name, frame.shape[1], frame.shape[0])
            k = cv2.waitKey(1)
            if len(faces_data) >= no_of_faces or k == 27:
                alive = False
                
        stop_camera(window_name)



        faces_data = np.asarray(faces_data)
        faces_data = faces_data.reshape(no_of_faces, -1)

        if "names.pkl" not in os.listdir("./data/"):
            names = [name] * no_of_faces
            with open("./data/names.pkl", "wb") as f:
                pickle.dump(names, f)
        else:
            with open("./data/names.pkl", "rb") as f:
                names = pickle.load(f)
            names = names + [name] * no_of_faces
            with open("./data/names.pkl", "wb") as f:
                pickle.dump(names, f)


        if "faces_data.pkl" not in os.listdir("./data/"):
            with open("./data/faces_data.pkl", "wb") as f:
                pickle.dump(faces_data, f)
        else:
            with open("./data/faces_data.pkl", "rb") as f:
                faces = pickle.load(f)
            faces = np.append(faces, faces_data, axis=0)
            with open("./data/faces_data.pkl", "wb") as f:
                pickle.dump(faces, f)
        print("Data saved successfully.")
        return
    
    if purpose == "recognize":

        alive = True
        
        #load the model
        with open("./data/knn.pkl", "rb") as f:
            knn_model = pickle.load(f)


        # Specify the CSV file name
        csv_file = "students.csv"

        # Initialize an empty dictionary
        students_dict = {}

        # Read the data from the CSV file into the dictionary
        with open(csv_file, mode='r', newline='') as file:
            reader = csv.DictReader(file)
            for row in reader:
                # Assuming the first column is the key and the second column is the value
                key = int(row["ID"])  # Convert the key to an integer
                value = row["Name"]
                students_dict[key] = value


        present_time = datetime.now().strftime("%I:%M:%S")
        year = datetime.now().year
        title = datetime.now().strftime("%B_%Y")
        today = datetime.now().strftime("%d-%m-%Y")

        #Try to load the attendance workbook
        try:
            wb = load_workbook(f"attendance_{datetime.now().year}.xlsx")
        #If it doesn't exist, create a new workbook
        except FileNotFoundError:
            wb = Workbook()
            wb.save(f"attendance_{datetime.now().year}.xlsx")
            wb = load_workbook(f"attendance_{datetime.now().year}.xlsx")

        #Check if the sheet for the current month exists, if not create it
        try:
            ws = wb[title]
        except KeyError:
            ws = wb.create_sheet(title, 0)
            first_day_of_month = datetime(year, datetime.now().month, 1)
            next_month = first_day_of_month.replace(month=first_day_of_month.month % 12 + 1, day=1)
            num_days = (next_month - first_day_of_month).days
            ws.cell(row=1, column=1, value="")
            ws.cell(row=1, column=2, value="")
            for i in range(1, 21):
                ws.cell(row=1, column=i + 2, value=i)

            for day in range(1, num_days + 1):
                date_str = (first_day_of_month + timedelta(days=day - 1)).strftime("%d-%m-%Y")
                ws.cell(row=day * 2, column=1, value=date_str)
                ws.cell(row=day * 2, column=2, value="am")
                ws.cell(row=day * 2 + 1, column=2, value="pm")
            wb.save(f"attendance_{year}.xlsx")


        while alive:
            # Read the frame
            ret, frame = video.read()
            if not ret:
                print("Error: Could not read frame.")
                break

            # Flip the frame
            frame = cv2.flip(frame, 1)

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY) 

            faces = face_detect.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4)

            # For each face, predict the name and mark attendance
            for (x, y, w, h) in faces:
                crop_face = frame[y:y+h, x:x+w, :]
                resized_face = cv2.resize(crop_face, (100, 100)).flatten().reshape
                name = knn_model.predict([resized_face])[0]
                cv2.putText(frame, name, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 1)
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 1)
                cv2.rectangle(frame, (x, y - 25), (x + w, y), (0, 255, 0), -1)

                for key in students_dict.keys():
                        if students_dict[key] == name:
                            roll_no = key
                            break
                    
                for col in ws.iter_cols(min_row = 1, max_col = 1, max_row = ws.max_row):
                    for cell in col:
                        if cell.value == today:
                            if datetime.now().hour < 12:
                                target_cell = ws.cell(row = cell.row, column = roll_no + 2)
                            else:
                                target_cell = ws.cell(row = cell.row + 1, column = roll_no + 2)                                

                            if target_cell.value is None:
                                target_cell.value = present_time
                                wb.save(f"attendance_{year}.xlsx")
                                print(f"Attendance marked for {name} at {present_time}")
                            else:
                                print(f"Attendance already marked for {name} at {target_cell.value}")
                            break

            cv2.imshow(window_name, frame)
            cv2.resizeWindow(window_name, frame.shape[1], frame.shape[0])

            k = cv2.waitKey(1)
            if k == 27:
                alive = False
        stop_camera(window_name)
        return

start_camera("add")